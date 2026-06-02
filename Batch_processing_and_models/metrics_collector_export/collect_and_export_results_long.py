"""
Collect metrics from results_long/ and export to metrics_results_long.xlsx.

Structure:
  reactor_<model>/<subfolder>/results.json   (itransformer, own, patchtst)
  reactor_<model>/results.json               (timellm — no subfolder)

Models: itransformer, own, patchtst, timellm
Per-batch metrics from per_experiment_metrics (overall/spike/non_spike).
Output: results_long_summary.json  +  metrics_results_long.xlsx
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE        = Path(__file__).parent / "results"
WANTED      = {"itransformer", "own", "patchtst", "timellm"}
BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU074", "KAU075", "KAU076", "KAU079"]

# --- styles ---
SECTION_FILL = PatternFill("solid", fgColor="2E4057")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="BBBBBB")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BASE_METRICS = [("mae", "MAE"), ("rmse", "RMSE"), ("mse_scaled", "MSE_scaled")]


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def pick_first_var(section: dict) -> dict:
    for vals in section.values():
        return vals
    return {}


def per_exp_to_entry(batch_data: dict) -> dict:
    return {
        "metrics":       {"nh4": pick_first_var(batch_data.get("overall",   {}))},
        "metrics_spike": {"nh4": pick_first_var(batch_data.get("spike",     {}))},
        "metrics_flat":  {"nh4": pick_first_var(batch_data.get("non_spike", {}))},
    }


def find_results_json(reactor_dir: Path) -> Path | None:
    """Return results.json path — either directly in reactor_dir or one level deeper."""
    direct = reactor_dir / "results.json"
    if direct.exists():
        return direct
    for child in reactor_dir.iterdir():
        if child.is_dir():
            candidate = child / "results.json"
            if candidate.exists():
                return candidate
    return None


def collect() -> dict:
    summary = {}
    for reactor_dir in sorted(BASE.iterdir()):
        if not reactor_dir.is_dir() or not reactor_dir.name.startswith("reactor_"):
            continue
        model = reactor_dir.name[len("reactor_"):]
        if model not in WANTED:
            continue

        rj = find_results_json(reactor_dir)
        if rj is None:
            print(f"  WARNING: no results.json found under {reactor_dir.name}")
            continue

        with open(rj) as f:
            d = json.load(f)

        per_exp = d.get("per_experiment_metrics", {})
        batches = {
            b: per_exp_to_entry(data)
            for b, data in per_exp.items()
            if b in BATCH_ORDER
        }
        if batches:
            summary[model] = {b: batches[b] for b in BATCH_ORDER if b in batches}

    return summary


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def flatten(summary: dict) -> list[dict]:
    rows = []
    for model, batches in summary.items():
        for batch in BATCH_ORDER:
            if batch not in batches:
                continue
            result = batches[batch]
            row = {"model": model, "batch": batch}
            for mkey in ("metrics", "metrics_spike", "metrics_flat"):
                if mkey not in result:
                    continue
                suffix = {"metrics": "", "metrics_spike": "_spike", "metrics_flat": "_flat"}[mkey]
                for var_vals in result[mkey].values():
                    for m, v in var_vals.items():
                        row[f"{m}{suffix}"] = v
            rows.append(row)
    return rows


def build_pivot(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if col not in df.columns:
        return pd.DataFrame()
    pivot = df.pivot_table(index="batch", columns="model", values=col, aggfunc="first")
    pivot = pivot.reindex([b for b in BATCH_ORDER if b in pivot.index])
    pivot.index.name = "Batch"
    return pivot


def set_column_widths(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (max(len(line) for line in str(c.value).splitlines()) for c in col_cells if c.value),
            default=10
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 30))


def write_block(ws, pivot: pd.DataFrame, start_row: int, label: str) -> int:
    n_cols = len(pivot.columns)
    cell = ws.cell(row=start_row, column=1, value=label)
    cell.font, cell.fill, cell.alignment = SECTION_FONT, SECTION_FILL, CENTER
    ws.row_dimensions[start_row].height = 22
    if n_cols > 0:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=n_cols + 1)
    hr = start_row + 1
    ws.cell(row=hr, column=1, value="Batch")
    for ci, col_name in enumerate(pivot.columns, 2):
        ws.cell(row=hr, column=ci, value=col_name)
    for cell in ws[hr]:
        if cell.column <= n_cols + 1:
            cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
    ws.row_dimensions[hr].height = 30
    for ri, (batch, row_data) in enumerate(pivot.iterrows(), 1):
        dr = hr + ri
        ws.cell(row=dr, column=1, value=batch)
        for ci, val in enumerate(row_data, 2):
            ws.cell(row=dr, column=ci, value=round(val, 6) if pd.notna(val) else None)
        for cell in ws[dr]:
            if cell.column <= n_cols + 1:
                cell.font, cell.alignment, cell.border = BODY_FONT, CENTER, BORDER
                if ri % 2 == 0:
                    cell.fill = ALT_FILL
        ws.row_dimensions[dr].height = 18
    return start_row + 1 + len(pivot) + 1


def export(flat_df: pd.DataFrame, out_path: Path):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for field, sheet_name in BASE_METRICS:
            overall = build_pivot(flat_df, field)
            spike   = build_pivot(flat_df, f"{field}_spike")
            flat_p  = build_pivot(flat_df, f"{field}_flat")
            if overall.empty:
                continue
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            cursor = 1
            cursor = write_block(ws, overall, cursor, f"{sheet_name}  —  Overall")
            cursor += 1
            cursor = write_block(ws, spike,   cursor, f"{sheet_name}  —  Spike")
            cursor += 1
            cursor = write_block(ws, flat_p,  cursor, f"{sheet_name}  —  Flat")
            set_column_widths(ws)

        flat_df.to_excel(writer, sheet_name="All Data", index=False)
        ws = writer.sheets["All Data"]
        for ri, row in enumerate(ws.iter_rows(max_row=len(flat_df)+1, max_col=len(flat_df.columns))):
            for cell in row:
                cell.alignment, cell.border = CENTER, BORDER
                if ri == 0:
                    cell.font, cell.fill = HEADER_FONT, HEADER_FILL
                else:
                    cell.font = BODY_FONT
                    if ri % 2 == 0:
                        cell.fill = ALT_FILL
        set_column_widths(ws)
        ws.row_dimensions[1].height = 30
    print(f"Written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    summary = collect()

    out_dir = Path(__file__).parent
    with open(out_dir / "results_long_summary.json", "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Written: {out_dir / 'results_long_summary.json'}")
    for model, batches in summary.items():
        print(f"  {model}: {list(batches.keys())}")

    rows = flatten(summary)
    if rows:
        export(pd.DataFrame(rows), out_dir / "metrics_results_long.xlsx")


if __name__ == "__main__":
    main()
