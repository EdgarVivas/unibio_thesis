"""
collect_and_export_mambats.py

Collects all MambATS metrics from results/reactor_mambats/ and writes
a single Excel file with one sheet per metric (MAE, RMSE, MSE_scaled).

Rows = batches, Columns = configurations, Sections = Overall / Spike / Non-spike.

Sources collected:
  absolute/       per-batch subfolders, per_experiment_metrics
  anchor/         per-batch subfolders, per_experiment_metrics
  concat_4batch   KAU084_KAU081_KAU071_KAU079/results.json
  no_token_7batch no_tokenKAU084_.../results.json

Usage:
    python collect_and_export_mambats.py
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent / "results" / "reactor_mambats"
OUT  = Path(__file__).parent / "metrics_mambats.xlsx"

BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU079",
               "KAU074", "KAU075", "KAU076"]

METRICS = ["mae", "rmse", "mse_scaled"]
METRIC_LABELS = {"mae": "MAE", "rmse": "RMSE", "mse_scaled": "MSE_scaled"}

CONFIG_ORDER = [
    "absolute",
    "anchor",
    "concat_4batch",
    "no_token_7batch",
]

FILL_SECTION = PatternFill("solid", fgColor="2E4057")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_ROW_A   = PatternFill("solid", fgColor="D6E4F0")
FILL_ROW_B   = PatternFill("solid", fgColor="FFFFFF")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD       = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _first_target(section_dict: dict) -> dict:
    if not section_dict:
        return {}
    return next(iter(section_dict.values()), {})


def _entry_from_pem(pem_entry: dict) -> dict:
    return {
        "overall": _first_target(pem_entry.get("overall", {})),
        "spike":   _first_target(pem_entry.get("spike", {})),
        "flat":    _first_target(pem_entry.get("non_spike", {})),
    }


def _entry_from_toplevel(d: dict) -> dict:
    return {
        "overall": _first_target(d.get("metrics", {})),
        "spike":   _first_target(d.get("metrics_spike", {})),
        "flat":    _first_target(d.get("metrics_flat", {})),
    }


def _load_json(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def collect_all() -> dict:
    data: dict = {}

    # ── per-batch folders: absolute, anchor ────────────────────────────────
    for config in ("absolute", "anchor"):
        folder = BASE / config
        if not folder.exists():
            print(f"  WARNING: {folder} not found — skipping")
            continue
        for batch_dir in sorted(folder.iterdir()):
            if not batch_dir.is_dir():
                continue
            batch = batch_dir.name
            rj = batch_dir / "results.json"
            if not rj.exists():
                continue
            d = _load_json(rj)
            pem = d.get("per_experiment_metrics", {})
            if batch in pem:
                entry = _entry_from_pem(pem[batch])
            else:
                entry = _entry_from_toplevel(d)
            data.setdefault(config, {})[batch] = entry
        found = sorted(data.get(config, {}).keys())
        print(f"  {config:<15}: {found}")

    # ── multi-batch concat files ─────────────────────────────────────────────
    concat_specs = {
        "concat_4batch":   "KAU084_KAU081_KAU071_KAU079",
        "no_token_7batch": "no_tokenKAU084_KAU081_KAU071_KAU074_KAU075_KAU076_KAU079",
    }
    for config, folder_name in concat_specs.items():
        rj = BASE / folder_name / "results.json"
        if not rj.exists():
            print(f"  WARNING: {rj} not found — skipping {config}")
            continue
        d = _load_json(rj)
        pem = d.get("per_experiment_metrics", {})
        for batch, pem_entry in pem.items():
            data.setdefault(config, {})[batch] = _entry_from_pem(pem_entry)
        found = sorted(data.get(config, {}).keys())
        print(f"  {config:<15}: {found}")

    return data


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def _style_cell(cell, fill=None, font=None, bold=False, align="center"):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    elif bold:
        cell.font = FONT_BOLD
    cell.alignment = Alignment(horizontal=align, vertical="center")


def write_excel(data: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_cols = 1 + len(CONFIG_ORDER)

    sections = [
        ("overall", "Overall"),
        ("spike",   "Spike"),
        ("flat",    "Non-Spike"),
    ]

    for metric in METRICS:
        ws = wb.create_sheet(title=METRIC_LABELS[metric])
        row = 1

        for section_key, section_label in sections:
            cell = ws.cell(row, 1, section_label)
            _style_cell(cell, fill=FILL_SECTION, font=FONT_WHITE_BOLD, align="left")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=total_cols)
            row += 1

            cell = ws.cell(row, 1, "Batch")
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            for ci, config in enumerate(CONFIG_ORDER, start=2):
                cell = ws.cell(row, ci, config)
                _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            row += 1

            active_batches = [
                b for b in BATCH_ORDER
                if any(
                    data.get(c, {}).get(b, {}).get(section_key, {}).get(metric) is not None
                    for c in CONFIG_ORDER
                )
            ]

            for ri, batch in enumerate(active_batches):
                fill = FILL_ROW_A if ri % 2 == 0 else FILL_ROW_B
                cell = ws.cell(row, 1, batch)
                _style_cell(cell, fill=fill, bold=True, align="left")
                for ci, config in enumerate(CONFIG_ORDER, start=2):
                    val = (data.get(config, {})
                               .get(batch, {})
                               .get(section_key, {})
                               .get(metric))
                    cell = ws.cell(row, ci)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val is not None:
                        cell.value = round(val, 6)
                        cell.number_format = "0.000000"
                row += 1

            row += 1

        ws.column_dimensions["A"].width = 12
        for ci, config in enumerate(CONFIG_ORDER, start=2):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(config) + 2, 14)

    wb.save(OUT)
    print(f"\nWritten: {OUT}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"\nCollecting from: {BASE}")
    print("-" * 60)
    data = collect_all()

    configs_found = [c for c in CONFIG_ORDER if c in data]
    all_batches = {b for cd in data.values() for b in cd}
    print(f"\nConfigs collected: {configs_found}")
    print(f"Writing {len(METRICS)} sheets × {len(all_batches)} active batches ...")
    write_excel(data)


if __name__ == "__main__":
    main()
