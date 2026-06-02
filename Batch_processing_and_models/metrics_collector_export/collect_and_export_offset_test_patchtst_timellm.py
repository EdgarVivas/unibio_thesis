"""
collect_and_export_offset_test_patchtst_timellm.py

Collects variable-ablation (offset_test) metrics for PatchTST and TimeLLM
from results/reactor_patchtst/offset_test/ and results/reactor_timellm/offset_test/.

Produces two Excel files:
  offset_test_patchtst.xlsx
  offset_test_timellm.xlsx

Each file: one sheet per metric (MAE, RMSE, MSE_scaled),
rows = batches, columns = variable configs, sections = Overall / Spike / Non-Spike.

Usage:
    python collect_and_export_offset_test_patchtst_timellm.py
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent

MODELS = {
    "patchtst": HERE / "results" / "reactor_patchtst" / "offset_test",
    "timellm":  HERE / "results" / "reactor_timellm"  / "offset_test",
}

BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU079"]

CONFIG_ORDER = [
    "nh4_only",
    "nh4_nh3",
    "nh4_no3_nh3",
    "nh4_no3_nh3_ch4",
    "nh4_no3_nh3_ch4_K",
    "all_without_no3",
]

CONFIG_ALIASES = {
    "all_withut_no3": "all_without_no3",
}

METRICS = ["mae", "rmse", "mse_scaled"]
METRIC_LABELS = {"mae": "MAE", "rmse": "RMSE", "mse_scaled": "MSE_scaled"}

FILL_SECTION = PatternFill("solid", fgColor="2E4057")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_ROW_A   = PatternFill("solid", fgColor="D6E4F0")
FILL_ROW_B   = PatternFill("solid", fgColor="FFFFFF")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD       = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _first_target(section_dict: dict) -> dict:
    if not section_dict:
        return {}
    return next(iter(section_dict.values()), {})


def _entry_from_pem(pem_entry: dict) -> dict:
    return {
        "overall": _first_target(pem_entry.get("overall", {})),
        "spike":   _first_target(pem_entry.get("spike", {})),
        "flat":    _first_target(pem_entry.get("non_spike", {})),
    }


def _entry_from_toplevel(d: dict) -> dict:
    return {
        "overall": _first_target(d.get("metrics", {})),
        "spike":   _first_target(d.get("metrics_spike", {})),
        "flat":    _first_target(d.get("metrics_flat", {})),
    }


def _load_json(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


def _normalise_config(name: str) -> str:
    return CONFIG_ALIASES.get(name, name)


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def collect_model(base: Path) -> dict:
    """Return data[config][batch] = {overall, spike, flat}."""
    data: dict = {}

    for config_dir in sorted(base.iterdir()):
        if not config_dir.is_dir():
            continue
        config = _normalise_config(config_dir.name)
        for batch_dir in sorted(config_dir.iterdir()):
            if not batch_dir.is_dir():
                continue
            batch = batch_dir.name
            rj = batch_dir / "results.json"
            if not rj.exists():
                continue
            d = _load_json(rj)
            pem = d.get("per_experiment_metrics", {})
            if batch in pem:
                entry = _entry_from_pem(pem[batch])
            else:
                entry = _entry_from_toplevel(d)
            data.setdefault(config, {})[batch] = entry

    for config in sorted(data):
        found = sorted(data[config].keys())
        print(f"  {config:<20}: {found}")

    return data


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def _style_cell(cell, fill=None, font=None, bold=False, align="center"):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    elif bold:
        cell.font = FONT_BOLD
    cell.alignment = Alignment(horizontal=align, vertical="center")


def write_excel(data: dict, out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Only include configs actually present
    configs = [c for c in CONFIG_ORDER if c in data]
    total_cols = 1 + len(configs)

    sections = [
        ("overall", "Overall"),
        ("spike",   "Spike"),
        ("flat",    "Non-Spike"),
    ]

    for metric in METRICS:
        ws = wb.create_sheet(title=METRIC_LABELS[metric])
        row = 1

        for section_key, section_label in sections:
            cell = ws.cell(row, 1, section_label)
            _style_cell(cell, fill=FILL_SECTION, font=FONT_WHITE_BOLD, align="left")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=total_cols)
            row += 1

            cell = ws.cell(row, 1, "Batch")
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            for ci, config in enumerate(configs, start=2):
                cell = ws.cell(row, ci, config)
                _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            row += 1

            active_batches = [
                b for b in BATCH_ORDER
                if any(
                    data.get(c, {}).get(b, {}).get(section_key, {}).get(metric) is not None
                    for c in configs
                )
            ]

            for ri, batch in enumerate(active_batches):
                fill = FILL_ROW_A if ri % 2 == 0 else FILL_ROW_B
                cell = ws.cell(row, 1, batch)
                _style_cell(cell, fill=fill, bold=True, align="left")
                for ci, config in enumerate(configs, start=2):
                    val = (data.get(config, {})
                               .get(batch, {})
                               .get(section_key, {})
                               .get(metric))
                    cell = ws.cell(row, ci)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val is not None:
                        cell.value = round(val, 6)
                        cell.number_format = "0.000000"
                row += 1

            row += 1

        ws.column_dimensions["A"].width = 12
        for ci, config in enumerate(configs, start=2):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(config) + 2, 14)

    wb.save(out_path)
    print(f"  Written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    for model_name, base in MODELS.items():
        out = HERE / f"offset_test_{model_name}.xlsx"
        print(f"\n{'='*60}")
        print(f"  {model_name.upper()}  —  {base}")
        print(f"{'='*60}")
        if not base.exists():
            print(f"  WARNING: folder not found — skipping")
            continue
        data = collect_model(base)
        write_excel(data, out)


if __name__ == "__main__":
    main()
