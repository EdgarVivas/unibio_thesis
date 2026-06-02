"""
collect_and_export_calibration_new_models.py

Collects calibration metrics for PatchTST, TimeLLM, and MambATS from their
respective results/reactor_* folders and writes one Excel file per model:
  calibration_patchtst.xlsx
  calibration_timellm.xlsx
  calibration_mambats.xlsx

Each file: sheets = 68% / 90% / 95%, sections = Overall / Spike,
rows = batches (incl. "all_batches" for aggregate runs), columns = configs.

Usage:
    python collect_and_export_calibration_new_models.py
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent

# Model definitions: name → (results folder, config list)
MODELS = {
    "patchtst": (
        HERE / "results" / "reactor_patchtst",
        ["absolute", "anchor", "one_by_one", "LOO", "finetune",
         "concat_4batch", "no_token_7batch", "token_7batch",
         "offset_nh4_only", "offset_nh4_nh3", "offset_nh4_no3_nh3",
         "offset_nh4_no3_nh3_ch4", "offset_nh4_no3_nh3_ch4_K", "offset_all_without_no3"],
    ),
    "timellm": (
        HERE / "results" / "reactor_timellm",
        ["absolute", "anchor", "one_by_one", "LOO", "finetune",
         "concat_4batch", "no_token_7batch", "token_7batch",
         "offset_nh4_only", "offset_nh4_nh3", "offset_nh4_no3_nh3",
         "offset_nh4_no3_nh3_ch4", "offset_nh4_no3_nh3_ch4_K", "offset_all_without_no3"],
    ),
    "mambats": (
        HERE / "results" / "reactor_mambats",
        ["absolute", "anchor", "concat_4batch", "no_token_7batch"],
    ),
}

# Folder names for multi-batch runs (map config key → subfolder name)
CONCAT_FOLDERS = {
    "concat_4batch":   "KAU084_KAU081_KAU071_KAU079",
    "no_token_7batch": "no_tokenKAU084_KAU081_KAU071_KAU074_KAU075_KAU076_KAU079",
    "token_7batch":    "token_KAU084_KAU081_KAU071_KAU074_KAU075_KAU076_KAU079",
}

BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU079",
               "KAU074", "KAU075", "KAU076", "all_batches"]

PCT_LEVELS = ["68%", "90%", "95%"]

FILL_SECTION = PatternFill("solid", fgColor="2E4057")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_ROW_A   = PatternFill("solid", fgColor="D6E4F0")
FILL_ROW_B   = PatternFill("solid", fgColor="FFFFFF")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD       = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_json(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


def _read_calib(d: dict) -> dict:
    """Extract {overall: {68%: v, ...}, spike: {68%: v, ...}} from a results dict."""
    def _norm(raw: dict) -> dict:
        # normalise keys like "68% (1σ)" → "68%"
        return {k.split(" (")[0].strip(): v for k, v in raw.items()} if raw else {}

    return {
        "overall": _norm(d.get("calibration", {})),
        "spike":   _norm(d.get("calibration_spike", {})),
    }


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def collect_model(base: Path, configs: list) -> dict:
    """Return data[config][batch] = {overall: {pct: v}, spike: {pct: v}}."""
    data: dict = {}

    for config in configs:
        if config in CONCAT_FOLDERS:
            # ── single results.json covering all batches ───────────────────
            folder_name = CONCAT_FOLDERS[config]
            rj = base / folder_name / "results.json"
            if not rj.exists():
                print(f"  WARNING: {rj} not found — skipping {config}")
                continue
            d = _load_json(rj)
            data.setdefault(config, {})["all_batches"] = _read_calib(d)
            print(f"  {config:<15}: [all_batches]")

        elif config == "LOO":
            # ── one results.json per held-out batch ────────────────────────
            loo_folder = base / "LOO"
            if not loo_folder.exists():
                print(f"  WARNING: {loo_folder} not found — skipping LOO")
                continue
            for loo_dir in sorted(loo_folder.iterdir()):
                if not loo_dir.is_dir():
                    continue
                rj = loo_dir / "results.json"
                if not rj.exists():
                    continue
                d = _load_json(rj)
                pem_keys = list(d.get("per_experiment_metrics", {}).keys())
                if not pem_keys:
                    continue
                batch = pem_keys[0]
                data.setdefault("LOO", {})[batch] = _read_calib(d)
            found = sorted(data.get("LOO", {}).keys())
            print(f"  {'LOO':<15}: {found}")

        elif config == "finetune":
            # ── test*/finetune_KAU* subfolders ─────────────────────────────
            ft_folder = base / "finetune"
            if not ft_folder.exists():
                print(f"  WARNING: {ft_folder} not found — skipping finetune")
                continue
            for test_dir in sorted(ft_folder.iterdir()):
                if not test_dir.is_dir():
                    continue
                for sub_dir in sorted(test_dir.iterdir()):
                    if not sub_dir.is_dir() or not sub_dir.name.startswith("finetune_"):
                        continue
                    rj = sub_dir / "results.json"
                    if not rj.exists():
                        continue
                    d = _load_json(rj)
                    batch = sub_dir.name[len("finetune_"):]
                    data.setdefault("finetune", {})[batch] = _read_calib(d)
            found = sorted(data.get("finetune", {}).keys())
            print(f"  {'finetune':<15}: {found}")

        elif config.startswith("offset_"):
            # ── offset_test/<var_config>/<batch>/results.json ──────────────
            var_config = config[len("offset_"):]
            folder = base / "offset_test" / var_config
            if not folder.exists():
                # try alias
                if var_config == "all_without_no3":
                    folder = base / "offset_test" / "all_withut_no3"
            if not folder.exists():
                print(f"  WARNING: {folder} not found — skipping {config}")
                continue
            for batch_dir in sorted(folder.iterdir()):
                if not batch_dir.is_dir():
                    continue
                batch = batch_dir.name
                rj = batch_dir / "results.json"
                if not rj.exists():
                    continue
                d = _load_json(rj)
                data.setdefault(config, {})[batch] = _read_calib(d)
            found = sorted(data.get(config, {}).keys())
            print(f"  {config:<25}: {found}")

        else:
            # ── per-batch subfolders (absolute, anchor, one_by_one) ────────
            folder = base / config
            if not folder.exists():
                print(f"  WARNING: {folder} not found — skipping {config}")
                continue
            for batch_dir in sorted(folder.iterdir()):
                if not batch_dir.is_dir():
                    continue
                batch = batch_dir.name
                rj = batch_dir / "results.json"
                if not rj.exists():
                    continue
                d = _load_json(rj)
                data.setdefault(config, {})[batch] = _read_calib(d)
            found = sorted(data.get(config, {}).keys())
            print(f"  {config:<15}: {found}")

    return data


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def _style_cell(cell, fill=None, font=None, bold=False, align="center"):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    elif bold:
        cell.font = FONT_BOLD
    cell.alignment = Alignment(horizontal=align, vertical="center")


def write_excel(data: dict, configs: list, out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    active_configs = [c for c in configs if c in data]
    total_cols = 1 + len(active_configs)

    sections = [
        ("overall", "Overall"),
        ("spike",   "Spike"),
    ]

    for pct in PCT_LEVELS:
        ws = wb.create_sheet(title=pct)
        row = 1

        for section_key, section_label in sections:
            # section title
            cell = ws.cell(row, 1, section_label)
            _style_cell(cell, fill=FILL_SECTION, font=FONT_WHITE_BOLD, align="left")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=total_cols)
            row += 1

            # column headers
            cell = ws.cell(row, 1, "Batch")
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            for ci, config in enumerate(active_configs, start=2):
                cell = ws.cell(row, ci, config)
                _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            row += 1

            # data rows — only batches that have at least one value
            active_batches = [
                b for b in BATCH_ORDER
                if any(
                    data.get(c, {}).get(b, {}).get(section_key, {}).get(pct) is not None
                    for c in active_configs
                )
            ]

            for ri, batch in enumerate(active_batches):
                fill = FILL_ROW_A if ri % 2 == 0 else FILL_ROW_B
                cell = ws.cell(row, 1, batch)
                _style_cell(cell, fill=fill, bold=True, align="left")
                for ci, config in enumerate(active_configs, start=2):
                    val = (data.get(config, {})
                               .get(batch, {})
                               .get(section_key, {})
                               .get(pct))
                    cell = ws.cell(row, ci)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val is not None:
                        cell.value = round(val, 4)
                        cell.number_format = "0.00%"
                row += 1

            row += 1  # blank between sections

        ws.column_dimensions["A"].width = 14
        for ci, config in enumerate(active_configs, start=2):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(config) + 2, 14)

    wb.save(out_path)
    print(f"  Written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    for model_name, (base, configs) in MODELS.items():
        out = HERE / f"calibration_{model_name}.xlsx"
        print(f"\n{'='*60}")
        print(f"  {model_name.upper()}  —  {base.name}")
        print(f"{'='*60}")
        if not base.exists():
            print(f"  WARNING: folder not found — skipping")
            continue
        data = collect_model(base, configs)
        write_excel(data, configs, out)


if __name__ == "__main__":
    main()
