"""
collect_and_export_timellm_transfer.py

Collects metrics and calibration for the TimeLLM transfer-learning runs
stored directly in results/reactor_timellm/KAU*/results.json (one file per batch).

Outputs:
  metrics_timellm_transfer.xlsx     — MAE / RMSE / MSE_scaled
  calibration_timellm_transfer.xlsx — 68% / 90% / 95%

Usage:
    python collect_and_export_timellm_transfer.py
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
BASE = HERE / "results" / "reactor_timellm"

BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU079",
               "KAU074", "KAU075", "KAU076"]

FILL_SECTION = PatternFill("solid", fgColor="2E4057")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_ROW_A   = PatternFill("solid", fgColor="D6E4F0")
FILL_ROW_B   = PatternFill("solid", fgColor="FFFFFF")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD       = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_json(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


def _first_target(d: dict) -> dict:
    return next(iter(d.values()), {}) if d else {}


def _style_cell(cell, fill=None, font=None, bold=False, align="center"):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    elif bold:
        cell.font = FONT_BOLD
    cell.alignment = Alignment(horizontal=align, vertical="center")


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def collect() -> dict:
    """Return {batch: {metrics: {overall,spike,flat}, calib: {overall,spike}}}."""
    data = {}
    for batch in BATCH_ORDER:
        rj = BASE / batch / "results.json"
        if not rj.exists():
            print(f"  WARNING: {rj} not found — skipping")
            continue
        d = _load_json(rj)
        data[batch] = {
            "metrics": {
                "overall": _first_target(d.get("metrics", {})),
                "spike":   _first_target(d.get("metrics_spike", {})),
                "flat":    _first_target(d.get("metrics_flat", {})),
            },
            "calib": {
                "overall": d.get("calibration", {}),
                "spike":   d.get("calibration_spike", {}),
            },
        }
        print(f"  {batch}: metrics={data[batch]['metrics']['overall']}  "
              f"calib={data[batch]['calib']['overall']}")
    return data


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------

def _write_metrics(data: dict, out: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    metrics    = ["mae", "rmse", "mse_scaled"]
    m_labels   = {"mae": "MAE", "rmse": "RMSE", "mse_scaled": "MSE_scaled"}
    sections   = [("overall", "Overall"), ("spike", "Spike"), ("flat", "Non-Spike")]
    col_label  = "transfer_learning"
    total_cols = 2  # batch + one value column

    for metric in metrics:
        ws = wb.create_sheet(title=m_labels[metric])
        row = 1

        for section_key, section_label in sections:
            cell = ws.cell(row, 1, section_label)
            _style_cell(cell, fill=FILL_SECTION, font=FONT_WHITE_BOLD, align="left")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=total_cols)
            row += 1

            cell = ws.cell(row, 1, "Batch")
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            cell = ws.cell(row, 2, col_label)
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            row += 1

            for ri, batch in enumerate(b for b in BATCH_ORDER if b in data):
                fill = FILL_ROW_A if ri % 2 == 0 else FILL_ROW_B
                cell = ws.cell(row, 1, batch)
                _style_cell(cell, fill=fill, bold=True, align="left")
                val = data[batch]["metrics"].get(section_key, {}).get(metric)
                cell = ws.cell(row, 2)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val is not None:
                    cell.value = round(val, 6)
                    cell.number_format = "0.000000"
                row += 1

            row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = max(len(col_label) + 2, 18)

    wb.save(out)
    print(f"\nWritten: {out}")


def _write_calibration(data: dict, out: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    pct_levels = ["68%", "90%", "95%"]
    sections   = [("overall", "Overall"), ("spike", "Spike")]
    col_label  = "transfer_learning"
    total_cols = 2

    for pct in pct_levels:
        ws = wb.create_sheet(title=pct)
        row = 1

        for section_key, section_label in sections:
            cell = ws.cell(row, 1, section_label)
            _style_cell(cell, fill=FILL_SECTION, font=FONT_WHITE_BOLD, align="left")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=total_cols)
            row += 1

            cell = ws.cell(row, 1, "Batch")
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            cell = ws.cell(row, 2, col_label)
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            row += 1

            for ri, batch in enumerate(b for b in BATCH_ORDER if b in data):
                fill = FILL_ROW_A if ri % 2 == 0 else FILL_ROW_B
                cell = ws.cell(row, 1, batch)
                _style_cell(cell, fill=fill, bold=True, align="left")
                val = data[batch]["calib"].get(section_key, {}).get(pct)
                cell = ws.cell(row, 2)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val is not None:
                    cell.value = round(val, 4)
                    cell.number_format = "0.00%"
                row += 1

            row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = max(len(col_label) + 2, 18)

    wb.save(out)
    print(f"Written: {out}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"\nCollecting from: {BASE}")
    print("-" * 60)
    data = collect()
    print(f"\nBatches found: {[b for b in BATCH_ORDER if b in data]}")

    _write_metrics(data,     HERE / "metrics_timellm_transfer.xlsx")
    _write_calibration(data, HERE / "calibration_timellm_transfer.xlsx")


if __name__ == "__main__":
    main()
