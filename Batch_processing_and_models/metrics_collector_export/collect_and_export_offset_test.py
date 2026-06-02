"""
Collect and export the variable ablation (offset_test) results.
Source: results_yesyes/offset_test/<model>/<reactor_<model>_<config>>/<batch>/results.json

Models and their folder names:
  itransformer   → itransformer  (reactor prefix: reactor_itransformer_)
  patchtst       → patchtst      (reactor prefix: reactor_patchtst_)
  single_timexer → single_timexer (reactor prefix: reactor_probabilistic_)
  timellm        → timellm       (reactor prefix: reactor_timellm_)

One Excel file per model: offset_test_<model>.xlsx
Within each file: rows = batches, columns = configs, sheets = MAE / RMSE / MSE_scaled
                  each sheet stacks Overall → Spike → Flat
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE        = Path(__file__).parent / "results_yesyes" / "offset_test"
BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU079"]
CONFIG_ORDER = [
    "nh4_only",
    "nh4_nh3",
    "nh4_no3_nh3",
    "nh4_no3_nh3_ch4",
    "nh4_no3_nh3_ch4_K",
    "all_without_no3",
]

MODELS = {
    "itransformer":   "reactor_itransformer_",
    "patchtst":       "reactor_patchtst_",
    "single_timexer": "reactor_probabilistic_",
    "timellm":        "reactor_timellm_",
}

# --- styles ---
SECTION_FILL = PatternFill("solid", fgColor="2E4057")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="BBBBBB")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BASE_METRICS = [("mae", "MAE"), ("rmse", "RMSE"), ("mse_scaled", "MSE_scaled")]


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def load_metrics(path: Path) -> dict | None:
    if not path.exists():
        return None
    with open(path) as f:
        d = json.load(f)
    entry = {}
    for key in ("metrics", "metrics_spike", "metrics_flat"):
        if key in d:
            entry[key] = d[key]
    return entry if entry else None


def normalise_config(raw_suffix: str) -> str:
    """Map any known variant spelling to the canonical CONFIG_ORDER label."""
    # handle typo: all_withut_no3 → all_without_no3
    return raw_suffix.replace("all_withut_no3", "all_without_no3")


def collect_model(model_dir: Path, reactor_prefix: str) -> dict:
    """
    Returns: { config: { batch: { metrics, metrics_spike, metrics_flat } } }
    """
    data = {}
    for config_dir in sorted(model_dir.iterdir()):
        if not config_dir.is_dir():
            continue
        name = config_dir.name
        if not name.startswith(reactor_prefix):
            continue
        raw_suffix = name[len(reactor_prefix):]
        config = normalise_config(raw_suffix)

        batches = {}
        for batch_dir in config_dir.iterdir():
            if batch_dir.is_dir() and batch_dir.name in BATCH_ORDER:
                entry = load_metrics(batch_dir / "results.json")
                if entry:
                    batches[batch_dir.name] = entry

        if batches:
            data[config] = {b: batches[b] for b in BATCH_ORDER if b in batches}

    return data


# ---------------------------------------------------------------------------
# Flatten
# ---------------------------------------------------------------------------

def flatten(model_data: dict) -> list[dict]:
    rows = []
    for config in CONFIG_ORDER:
        if config not in model_data:
            continue
        for batch in BATCH_ORDER:
            if batch not in model_data[config]:
                continue
            result = model_data[config][batch]
            row = {"config": config, "batch": batch}
            for mkey in ("metrics", "metrics_spike", "metrics_flat"):
                if mkey not in result:
                    continue
                suffix = {"metrics": "", "metrics_spike": "_spike", "metrics_flat": "_flat"}[mkey]
                for var_vals in result[mkey].values():
                    for m, v in var_vals.items():
                        row[f"{m}{suffix}"] = v
            rows.append(row)
    return rows


def build_pivot(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if col not in df.columns:
        return pd.DataFrame()
    pivot = df.pivot_table(index="batch", columns="config", values=col, aggfunc="first")
    pivot = pivot.reindex([b for b in BATCH_ORDER if b in pivot.index])
    # order columns by CONFIG_ORDER
    ordered = [c for c in CONFIG_ORDER if c in pivot.columns]
    pivot = pivot[ordered]
    pivot.index.name = "Batch"
    return pivot


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def set_column_widths(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (max(len(line) for line in str(c.value).splitlines()) for c in col_cells if c.value),
            default=10
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 30))


def write_block(ws, pivot: pd.DataFrame, start_row: int, label: str) -> int:
    n_cols = len(pivot.columns)
    cell = ws.cell(row=start_row, column=1, value=label)
    cell.font, cell.fill, cell.alignment = SECTION_FONT, SECTION_FILL, CENTER
    ws.row_dimensions[start_row].height = 22
    if n_cols > 0:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=n_cols + 1)
    hr = start_row + 1
    ws.cell(row=hr, column=1, value="Batch")
    for ci, col_name in enumerate(pivot.columns, 2):
        ws.cell(row=hr, column=ci, value=col_name)
    for cell in ws[hr]:
        if cell.column <= n_cols + 1:
            cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
    ws.row_dimensions[hr].height = 30
    for ri, (batch, row_data) in enumerate(pivot.iterrows(), 1):
        dr = hr + ri
        ws.cell(row=dr, column=1, value=batch)
        for ci, val in enumerate(row_data, 2):
            ws.cell(row=dr, column=ci, value=round(val, 6) if pd.notna(val) else None)
        for cell in ws[dr]:
            if cell.column <= n_cols + 1:
                cell.font, cell.alignment, cell.border = BODY_FONT, CENTER, BORDER
                if ri % 2 == 0:
                    cell.fill = ALT_FILL
        ws.row_dimensions[dr].height = 18
    return start_row + 1 + len(pivot) + 1


def export_model(flat_df: pd.DataFrame, model_name: str, out_path: Path):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for field, sheet_name in BASE_METRICS:
            overall = build_pivot(flat_df, field)
            spike   = build_pivot(flat_df, f"{field}_spike")
            flat_p  = build_pivot(flat_df, f"{field}_flat")
            if overall.empty:
                continue
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            cursor = 1
            cursor = write_block(ws, overall, cursor, f"{sheet_name}  —  Overall")
            cursor += 1
            cursor = write_block(ws, spike,   cursor, f"{sheet_name}  —  Spike")
            cursor += 1
            cursor = write_block(ws, flat_p,  cursor, f"{sheet_name}  —  Flat")
            set_column_widths(ws)

        flat_df.to_excel(writer, sheet_name="All Data", index=False)
        ws = writer.sheets["All Data"]
        for ri, row in enumerate(ws.iter_rows(max_row=len(flat_df)+1, max_col=len(flat_df.columns))):
            for cell in row:
                cell.alignment, cell.border = CENTER, BORDER
                if ri == 0:
                    cell.font, cell.fill = HEADER_FONT, HEADER_FILL
                else:
                    cell.font = BODY_FONT
                    if ri % 2 == 0:
                        cell.fill = ALT_FILL
        set_column_widths(ws)
        ws.row_dimensions[1].height = 30
    print(f"Written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    out_dir = Path(__file__).parent

    for model, reactor_prefix in MODELS.items():
        model_dir = BASE / model
        if not model_dir.exists():
            print(f"  WARNING: {model_dir} not found, skipping")
            continue

        model_data = collect_model(model_dir, reactor_prefix)
        print(f"\n{model}:")
        for config, batches in model_data.items():
            print(f"  {config}: {list(batches.keys())}")

        rows = flatten(model_data)
        if rows:
            export_model(
                pd.DataFrame(rows),
                model,
                out_dir / f"offset_test_{model}.xlsx"
            )


if __name__ == "__main__":
    main()
