"""
collect_and_export_calibration.py

Collects calibration coverage (68% / 90% / 95%) from all result folders.
Calibration = fraction of true targets that fall within the predicted
confidence interval (ideal: 68% ≈ 0.68, 90% ≈ 0.90, 95% ≈ 0.95).

Outputs (written next to this file):
  calibration_anchor.xlsx          — results_yes anchor    (per batch)
  calibration_absolute.xlsx        — results_yes absolute  (per batch)
  calibration_modalities.xlsx      — results_yesyes modalities
  calibration_concat.xlsx          — results_concat        (aggregate per model)
  calibration_results_long.xlsx    — results_long          (aggregate per model)
  calibration_token.xlsx           — results_token         (aggregate per model)
  calibration_offset_<model>.xlsx  — results_yesyes/offset_test (per batch × config) × 4

Notes on aggregate vs per-batch:
  - Old-format files (one results.json per batch) → calibration IS per-batch.
  - New-format files (one results.json covering all batches via per_experiment_metrics)
    → calibration is aggregate across all batches; shown as a single "all_batches" row.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE        = Path(__file__).parent
BASE_YES    = HERE / "results_yes"
BASE_YESYES = HERE / "results_yesyes"
BASE_CONCAT = HERE / "results_concat"
BASE_LONG   = HERE / "results_long"
BASE_TOKEN  = HERE / "results_token"

BATCH_4    = ["KAU084", "KAU081", "KAU071", "KAU079"]
BATCH_7    = ["KAU084", "KAU081", "KAU071", "KAU074", "KAU075", "KAU076", "KAU079"]
THRESHOLDS = ["68%", "90%", "95%"]

OFFSET_MODELS = {
    "itransformer":   "reactor_itransformer_",
    "patchtst":       "reactor_patchtst_",
    "single_timexer": "reactor_probabilistic_",
    "timellm":        "reactor_timellm_",
}
CONFIG_ORDER = [
    "nh4_only", "nh4_nh3", "nh4_no3_nh3",
    "nh4_no3_nh3_ch4", "nh4_no3_nh3_ch4_K", "all_without_no3",
]
MODALITY_MODELS = ["own_timexer", "patchtst", "itransformer", "timellm"]
MODALITY_ORDER  = ["transfer_learning", "one_by_one", "concatenated", "loo", "finetune"]
MODEL_ORDER_MOD = ["own_timexer", "patchtst", "itransformer", "timellm"]

# --- styles ---
SECTION_FILL = PatternFill("solid", fgColor="2E4057")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="BBBBBB")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Calibration loaders
# ---------------------------------------------------------------------------

def load_calib(path: Path) -> dict | None:
    """Read calibration/calibration_spike/calibration_flat from results.json."""
    if not path.exists():
        return None
    with open(path) as f:
        d = json.load(f)
    entry = {}
    for key in ("calibration", "calibration_spike", "calibration_flat"):
        if d.get(key):
            entry[key] = d[key]
    return entry or None


def calib_to_row_fields(entry: dict) -> dict:
    """Convert calibration entry to flat column key→value pairs.

    calibration      → c68, c90, c95
    calibration_spike → c68_spike, c90_spike, c95_spike
    calibration_flat  → c68_flat, c90_flat, c95_flat
    """
    fields = {}
    for ckey, suffix in [
        ("calibration",       ""),
        ("calibration_spike", "_spike"),
        ("calibration_flat",  "_flat"),
    ]:
        if ckey not in entry:
            continue
        for pct_label, val in entry[ckey].items():
            # normalise "68% (1σ)" / "90% (1.645σ)" → "68%" / "90%"
            clean = pct_label.split(" (")[0].strip()
            col = "c" + clean.replace("%", "") + suffix  # "c68", "c90", "c95", "c68_spike" ...
            fields[col] = val
    return fields


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def set_column_widths(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (max(len(line) for line in str(c.value).splitlines()) for c in col_cells if c.value),
            default=10,
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 30))


def write_block(ws, pivot: pd.DataFrame, start_row: int, label: str) -> int:
    n_cols = len(pivot.columns)
    cell = ws.cell(row=start_row, column=1, value=label)
    cell.font, cell.fill, cell.alignment = SECTION_FONT, SECTION_FILL, CENTER
    ws.row_dimensions[start_row].height = 22
    if n_cols > 0:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=n_cols + 1)
    hr = start_row + 1
    ws.cell(row=hr, column=1, value="Batch")
    for ci, col_name in enumerate(pivot.columns, 2):
        ws.cell(row=hr, column=ci, value=col_name)
    for cell in ws[hr]:
        if cell.column <= n_cols + 1:
            cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
    ws.row_dimensions[hr].height = 30
    for ri, (idx, row_data) in enumerate(pivot.iterrows(), 1):
        dr = hr + ri
        ws.cell(row=dr, column=1, value=idx)
        for ci, val in enumerate(row_data, 2):
            ws.cell(row=dr, column=ci, value=round(val, 6) if pd.notna(val) else None)
        for cell in ws[dr]:
            if cell.column <= n_cols + 1:
                cell.font, cell.alignment, cell.border = BODY_FONT, CENTER, BORDER
                if ri % 2 == 0:
                    cell.fill = ALT_FILL
        ws.row_dimensions[dr].height = 18
    return start_row + 1 + len(pivot) + 1


def export_to_excel(
    flat_df: pd.DataFrame,
    out_path: Path,
    row_col: str,
    col_col: str,
    row_order: list,
    col_order: list | None = None,
):
    """Write calibration Excel file.

    Each sheet = one threshold (68%, 90%, 95%).
    Within each sheet: Overall block then Spike block stacked.
    """
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for thr in THRESHOLDS:
            safe = "c" + thr.replace("%", "")      # "c68", "c90", "c95"
            overall_col = safe
            spike_col   = safe + "_spike"

            if overall_col not in flat_df.columns:
                continue

            def make_pivot(col):
                if col not in flat_df.columns:
                    return pd.DataFrame()
                piv = flat_df.pivot_table(
                    index=row_col, columns=col_col, values=col, aggfunc="first"
                )
                piv = piv.reindex([r for r in row_order if r in piv.index])
                if col_order:
                    piv = piv[[c for c in col_order if c in piv.columns]]
                piv.index.name = row_col.capitalize()
                return piv

            overall = make_pivot(overall_col)
            spike   = make_pivot(spike_col)

            pd.DataFrame().to_excel(writer, sheet_name=thr)
            ws = writer.sheets[thr]
            cursor = 1
            cursor = write_block(ws, overall, cursor, f"{thr}  —  Overall")
            cursor += 1
            cursor = write_block(ws, spike,   cursor, f"{thr}  —  Spike")
            set_column_widths(ws)

        flat_df.to_excel(writer, sheet_name="All Data", index=False)
        ws = writer.sheets["All Data"]
        for ri, row in enumerate(ws.iter_rows(max_row=len(flat_df) + 1, max_col=len(flat_df.columns))):
            for cell in row:
                cell.alignment, cell.border = CENTER, BORDER
                if ri == 0:
                    cell.font, cell.fill = HEADER_FONT, HEADER_FILL
                else:
                    cell.font = BODY_FONT
                    if ri % 2 == 0:
                        cell.fill = ALT_FILL
        set_column_widths(ws)
        ws.row_dimensions[1].height = 30

    print(f"Written: {out_path}")


# ---------------------------------------------------------------------------
# Collect — results_yes (anchor / absolute) — per batch
# ---------------------------------------------------------------------------

def _scan_batch_dirs(root: Path, batch_list: list, aliases: dict = {}, exclude: set = set()) -> dict:
    """Walk root (up to two levels) for batch folders; return {batch: calib_entry}."""
    batches = {}
    if not root.exists():
        return batches
    for child in sorted(root.iterdir()):
        if not child.is_dir() or child.name in exclude:
            continue
        name = aliases.get(child.name, child.name)
        if name in batch_list:
            entry = load_calib(child / "results.json")
            if entry:
                batches[name] = entry
        else:
            for sub in sorted(child.iterdir()):
                if not sub.is_dir():
                    continue
                name2 = aliases.get(sub.name, sub.name)
                if name2 in batch_list:
                    entry = load_calib(sub / "results.json")
                    if entry:
                        batches[name2] = entry
    return {b: batches[b] for b in batch_list if b in batches}


def collect_yes_mode(mode: str) -> dict:
    """Returns {model: {batch: calib_entry}} for mode ∈ {"anchor", "absolute"}."""
    EXCL_IT_ABS = {"reactor_itransformer_conservative_noresidual"}
    CSDI_ABS_ALIAS = {"KAU084_revin_all_var": "KAU084"}

    # Each entry: (root_path, aliases, exclude)
    MODEL_ROOTS = {
        "csdi": (
            (BASE_YES / "csdi" / "absolute" / "all_vars") if mode == "absolute"
            else (BASE_YES / "csdi" / "anchor"),
            CSDI_ABS_ALIAS if mode == "absolute" else {},
            set(),
        ),
        "itransformer": (
            BASE_YES / "itransformer" / mode,
            {},
            EXCL_IT_ABS if mode == "absolute" else set(),
        ),
        "mambats":     (BASE_YES / "mambats"    / mode,                         {}, set()),
        "ncde":        (BASE_YES / "ncde"        / mode,                         {}, set()),
        "own_timexer":     (BASE_YES / "own_timexer" / mode / "single_endo",       {}, set()),
        "own_timexer_two": (BASE_YES / "own_timexer" / mode / "twoendo",          {}, set()),
        "patchtst":        (BASE_YES / "patchtst"    / f"reactor_patchtst_{mode}",{}, set()),
        "tft":             (BASE_YES / "tft"         / mode,                      {}, set()),
        "timeXer":         (BASE_YES / "timeXer"     / mode,                      {}, set()),
        "timellm":         (BASE_YES / "timellm"     / mode,                      {}, set()),
        "reactor_csdi":    (BASE_YES / "reactor_csdi",                            {}, set()),
    }

    summary = {}
    for model, (root, aliases, excl) in MODEL_ROOTS.items():
        batches = _scan_batch_dirs(root, BATCH_4, aliases, excl)
        if batches:
            summary[model] = batches
    return summary


def flatten_yes(summary: dict) -> pd.DataFrame:
    rows = []
    for model, batches in summary.items():
        for batch, entry in batches.items():
            row = {"model": model, "batch": batch}
            row.update(calib_to_row_fields(entry))
            rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ---------------------------------------------------------------------------
# Collect — results_yesyes/offset_test — per batch × config
# ---------------------------------------------------------------------------

def normalise_config(raw: str) -> str:
    return raw.replace("all_withut_no3", "all_without_no3")


def collect_offset_model(model_dir: Path, reactor_prefix: str) -> dict:
    """Returns {config: {batch: calib_entry}}."""
    data = {}
    for config_dir in sorted(model_dir.iterdir()):
        if not config_dir.is_dir() or not config_dir.name.startswith(reactor_prefix):
            continue
        config = normalise_config(config_dir.name[len(reactor_prefix):])
        batches = {}
        for batch_dir in config_dir.iterdir():
            if batch_dir.is_dir() and batch_dir.name in BATCH_4:
                entry = load_calib(batch_dir / "results.json")
                if entry:
                    batches[batch_dir.name] = entry
        if batches:
            data[config] = {b: batches[b] for b in BATCH_4 if b in batches}
    return data


def flatten_offset(model_data: dict) -> pd.DataFrame:
    rows = []
    for config in CONFIG_ORDER:
        if config not in model_data:
            continue
        for batch in BATCH_4:
            if batch not in model_data[config]:
                continue
            row = {"config": config, "batch": batch}
            row.update(calib_to_row_fields(model_data[config][batch]))
            rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ---------------------------------------------------------------------------
# Collect — aggregate files (results_concat, results_long, results_token)
# ---------------------------------------------------------------------------

def collect_aggregate(base: Path, wanted: set) -> dict:
    """Returns {model: {"all_batches": calib_entry}} for aggregate files.

    Structure: base/reactor_<model>/(optional_subfolder/)results.json
    """
    data = {}
    for reactor_dir in sorted(base.iterdir()):
        if not reactor_dir.is_dir() or not reactor_dir.name.startswith("reactor_"):
            continue
        model = reactor_dir.name[len("reactor_"):]
        if model not in wanted:
            continue
        # Find results.json (direct or one level deep)
        rj = reactor_dir / "results.json"
        if not rj.exists():
            for child in reactor_dir.iterdir():
                if child.is_dir():
                    candidate = child / "results.json"
                    if candidate.exists():
                        rj = candidate
                        break
        entry = load_calib(rj)
        if entry:
            data[model] = {"all_batches": entry}
    return data


def flatten_aggregate(summary: dict) -> pd.DataFrame:
    rows = []
    for model, batches in summary.items():
        for batch, entry in batches.items():
            row = {"model": model, "batch": batch}
            row.update(calib_to_row_fields(entry))
            rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ---------------------------------------------------------------------------
# Collect — modalities (results_yesyes)
# ---------------------------------------------------------------------------

def _per_exp_calib(path: Path, batch: str) -> dict | None:
    """New-format files: calibration at top level, but this file covers ONE batch."""
    return load_calib(path)


def collect_modality_transfer() -> dict:
    """Transfer learning — old format, per batch."""
    ROOTS = {
        "own_timexer":  BASE_YES / "own_timexer"  / "absolute" / "single_endo",
        "patchtst":     BASE_YES / "patchtst"      / "reactor_patchtst_absolute",
        "itransformer": BASE_YES / "itransformer"  / "absolute" / "reactor_itransformer_noresidual",
        "timellm":      BASE_YES / "timellm"       / "absolute",
    }
    data = {}
    for model in MODALITY_MODELS:
        root = ROOTS.get(model)
        if not root or not root.exists():
            continue
        batches = _scan_batch_dirs(root, BATCH_4)
        if batches:
            data[model] = batches
    return data


def collect_modality_one_by_one() -> dict:
    """One-by-one — one results.json per batch folder → per-batch calibration."""
    data = {}
    base = BASE_YESYES / "one_by_one"
    for model in MODALITY_MODELS:
        model_dir = base / model
        if not model_dir.exists():
            continue
        batches = {}
        for child in model_dir.iterdir():
            if not child.is_dir():
                continue
            for batch_dir in child.iterdir():
                if batch_dir.is_dir() and batch_dir.name in BATCH_4:
                    entry = load_calib(batch_dir / "results.json")
                    if entry:
                        batches[batch_dir.name] = entry
        if batches:
            data[model] = {b: batches[b] for b in BATCH_4 if b in batches}
    return data


def collect_modality_concatenated() -> dict:
    """Concatenated — one file per model covering all batches → aggregate."""
    data = {}
    base = BASE_YESYES / "concatenated_test"
    for model in MODALITY_MODELS:
        model_dir = base / model
        if not model_dir.exists():
            continue
        if model == "timellm":
            rj = model_dir / "results.json"
        else:
            target = None
            for child in model_dir.iterdir():
                if child.is_dir() and child.name.lower().replace("_", "").startswith("notoken"):
                    target = child
                    break
            if target is None:
                continue
            rj = target / "results.json"
        entry = load_calib(rj)
        if entry:
            data[model] = {"all_batches": entry}
    return data


def collect_modality_loo() -> dict:
    """Leave-one-out — top-level calibration; batch identified from per_experiment_metrics key."""
    data = {}
    base = BASE_YESYES / "loo"
    for model in MODALITY_MODELS:
        model_dir = base / model
        if not model_dir.exists():
            continue
        batches = {}
        for child in model_dir.iterdir():
            if not child.is_dir() or not child.name.startswith("LOO_"):
                continue
            rj = child / "results.json"
            if not rj.exists():
                continue
            with open(rj) as f:
                d = json.load(f)
            per_exp = d.get("per_experiment_metrics", {})
            if not per_exp:
                continue
            batch = next(iter(per_exp))
            if batch not in BATCH_4:
                continue
            entry = load_calib(rj)
            if entry:
                batches[batch] = entry
        if batches:
            data[model] = {b: batches[b] for b in BATCH_4 if b in batches}
    return data


def collect_modality_finetune() -> dict:
    """Finetune — one file per batch (via per_experiment_metrics key)."""
    data = {}
    base = BASE_YESYES / "finetune"
    for model in MODALITY_MODELS:
        model_dir = base / model
        if not model_dir.exists():
            continue
        batches = {}

        def _scan(tests_root: Path):
            for test_dir in tests_root.iterdir():
                if not test_dir.is_dir() or not test_dir.name.startswith("test"):
                    continue
                for sub in test_dir.iterdir():
                    if not sub.is_dir() or not sub.name.startswith("finetune_"):
                        continue
                    rj = sub / "results.json"
                    if not rj.exists():
                        continue
                    with open(rj) as f:
                        d = json.load(f)
                    per_exp = d.get("per_experiment_metrics", {})
                    for batch in per_exp:
                        if batch in BATCH_4:
                            entry = load_calib(rj)
                            if entry:
                                batches[batch] = entry

        if model == "timellm":
            _scan(model_dir)
        else:
            for child in model_dir.iterdir():
                if child.is_dir() and child.name.startswith("reactor_"):
                    _scan(child)

        if batches:
            data[model] = {b: batches[b] for b in BATCH_4 if b in batches}
    return data


def flatten_modalities(summary: dict) -> pd.DataFrame:
    rows = []
    for modality, models in summary.items():
        for model, batches in models.items():
            for batch, entry in batches.items():
                row = {"modality": modality, "model": model, "batch": batch}
                row.update(calib_to_row_fields(entry))
                rows.append(row)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def export_modalities_excel(flat_df: pd.DataFrame, out_path: Path):
    """Export modalities calibration: cols = model+modality, rows = batches."""
    all_batches = BATCH_4 + ["all_batches"]

    def make_col_label(r):
        return f"{r['model']}\n({r['modality']})"

    df = flat_df.copy()
    df["col"] = df.apply(make_col_label, axis=1)

    # Determine column order: group by model, then modality
    ordered_cols = []
    for model in MODEL_ORDER_MOD:
        for mod in MODALITY_ORDER:
            col_name = f"{model}\n({mod})"
            if col_name in df["col"].values:
                ordered_cols.append(col_name)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for thr in THRESHOLDS:
            safe = "c" + thr.replace("%", "")
            overall_col, spike_col = safe, safe + "_spike"
            if overall_col not in df.columns:
                continue

            def make_pivot(col):
                if col not in df.columns:
                    return pd.DataFrame()
                piv = df.pivot_table(index="batch", columns="col", values=col, aggfunc="first")
                piv = piv.reindex([b for b in all_batches if b in piv.index])
                ordered = [c for c in ordered_cols if c in piv.columns]
                piv = piv[ordered]
                piv.index.name = "Batch"
                return piv

            overall = make_pivot(overall_col)
            spike   = make_pivot(spike_col)

            pd.DataFrame().to_excel(writer, sheet_name=thr)
            ws = writer.sheets[thr]
            cursor = 1
            cursor = write_block(ws, overall, cursor, f"{thr}  —  Overall")
            cursor += 1
            cursor = write_block(ws, spike,   cursor, f"{thr}  —  Spike")
            set_column_widths(ws)

        flat_df.to_excel(writer, sheet_name="All Data", index=False)
        ws = writer.sheets["All Data"]
        for ri, row in enumerate(ws.iter_rows(max_row=len(flat_df) + 1, max_col=len(flat_df.columns))):
            for cell in row:
                cell.alignment, cell.border = CENTER, BORDER
                if ri == 0:
                    cell.font, cell.fill = HEADER_FONT, HEADER_FILL
                else:
                    cell.font = BODY_FONT
                    if ri % 2 == 0:
                        cell.fill = ALT_FILL
        set_column_widths(ws)
        ws.row_dimensions[1].height = 30

    print(f"Written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    out = HERE

    # 1. results_yes anchor
    print("\n=== calibration_anchor ===")
    data = collect_yes_mode("anchor")
    for model, batches in data.items():
        print(f"  {model}: {list(batches.keys())}")
    df = flatten_yes(data)
    if not df.empty:
        export_to_excel(df, out / "calibration_anchor.xlsx",
                        row_col="batch", col_col="model",
                        row_order=BATCH_4, col_order=list(data.keys()))

    # 2. results_yes absolute
    print("\n=== calibration_absolute ===")
    data = collect_yes_mode("absolute")
    for model, batches in data.items():
        print(f"  {model}: {list(batches.keys())}")
    df = flatten_yes(data)
    if not df.empty:
        export_to_excel(df, out / "calibration_absolute.xlsx",
                        row_col="batch", col_col="model",
                        row_order=BATCH_4, col_order=list(data.keys()))

    # 3. Modalities (results_yesyes)
    print("\n=== calibration_modalities ===")
    modality_summary = {
        "transfer_learning": collect_modality_transfer(),
        "one_by_one":        collect_modality_one_by_one(),
        "concatenated":      collect_modality_concatenated(),
        "loo":               collect_modality_loo(),
        "finetune":          collect_modality_finetune(),
    }
    for modality, models in modality_summary.items():
        for model, batches in models.items():
            print(f"  {modality}/{model}: {list(batches.keys())}")
    df = flatten_modalities(modality_summary)
    if not df.empty:
        export_modalities_excel(df, out / "calibration_modalities.xlsx")

    # 4. results_concat
    print("\n=== calibration_concat ===")
    data = collect_aggregate(BASE_CONCAT, {"csdi", "mambats", "ncde", "probabilistic", "tft", "timexer"})
    for model, batches in data.items():
        print(f"  {model}: {list(batches.keys())}")
    df = flatten_aggregate(data)
    if not df.empty:
        export_to_excel(df, out / "calibration_concat.xlsx",
                        row_col="batch", col_col="model",
                        row_order=["all_batches"], col_order=list(data.keys()))

    # 5. results_long
    print("\n=== calibration_results_long ===")
    data = collect_aggregate(BASE_LONG, {"itransformer", "own", "patchtst", "timellm"})
    for model, batches in data.items():
        print(f"  {model}: {list(batches.keys())}")
    df = flatten_aggregate(data)
    if not df.empty:
        export_to_excel(df, out / "calibration_results_long.xlsx",
                        row_col="batch", col_col="model",
                        row_order=["all_batches"], col_order=list(data.keys()))

    # 6. results_token
    print("\n=== calibration_token ===")
    data = collect_aggregate(BASE_TOKEN, {"itransformer", "own", "patchtst", "timellm"})
    for model, batches in data.items():
        print(f"  {model}: {list(batches.keys())}")
    df = flatten_aggregate(data)
    if not df.empty:
        export_to_excel(df, out / "calibration_token.xlsx",
                        row_col="batch", col_col="model",
                        row_order=["all_batches"], col_order=list(data.keys()))

    # 7. Offset test (results_yesyes/offset_test) — one file per model
    print("\n=== calibration_offset_test ===")
    offset_base = BASE_YESYES / "offset_test"
    for model, prefix in OFFSET_MODELS.items():
        model_dir = offset_base / model
        if not model_dir.exists():
            print(f"  WARNING: {model_dir} not found, skipping")
            continue
        model_data = collect_offset_model(model_dir, prefix)
        print(f"\n  {model}:")
        for config, batches in model_data.items():
            print(f"    {config}: {list(batches.keys())}")
        df = flatten_offset(model_data)
        if not df.empty:
            export_to_excel(df, out / f"calibration_offset_{model}.xlsx",
                            row_col="batch", col_col="config",
                            row_order=BATCH_4, col_order=CONFIG_ORDER)


if __name__ == "__main__":
    main()
