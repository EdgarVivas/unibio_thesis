"""
collect_and_export_patchtst.py

Collects all PatchTST metrics from results/reactor_patchtst/ and writes
a single Excel file with one sheet per metric (MAE, RMSE, MSE_scaled).

Rows = batches, Columns = configurations, Sections = Overall / Spike / Non-spike.

Sources collected:
  absolute/       per-batch subfolders, per_experiment_metrics
  anchor/         per-batch subfolders, per_experiment_metrics
  one_by_one/     per-batch subfolders, per_experiment_metrics
  LOO/            per-LOO folder, top-level metrics (batch from pem key)
  finetune/       test* / finetune_KAU* subfolders, per_experiment_metrics
  concat_4batch   KAU084_KAU081_KAU071_KAU079/results.json
  no_token_7batch no_tokenKAU084_.../results.json
  token_7batch    token_KAU084_.../results.json

Usage:
    python collect_and_export_patchtst.py
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent / "results" / "reactor_patchtst"
OUT  = Path(__file__).parent / "metrics_patchtst.xlsx"

BATCH_ORDER = ["KAU084", "KAU081", "KAU071", "KAU079",
               "KAU074", "KAU075", "KAU076"]

METRICS = ["mae", "rmse", "mse_scaled"]
METRIC_LABELS = {"mae": "MAE", "rmse": "RMSE", "mse_scaled": "MSE_scaled"}

CONFIG_ORDER = [
    "absolute",
    "anchor",
    "one_by_one",
    "LOO",
    "finetune",
    "concat_4batch",
    "no_token_7batch",
    "token_7batch",
]

FILL_SECTION = PatternFill("solid", fgColor="2E4057")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_ROW_A   = PatternFill("solid", fgColor="D6E4F0")
FILL_ROW_B   = PatternFill("solid", fgColor="FFFFFF")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD       = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _first_target(section_dict: dict) -> dict:
    """Extract metrics from {target: {mae, rmse, mse_scaled}} → {mae, rmse, mse_scaled}."""
    if not section_dict:
        return {}
    return next(iter(section_dict.values()), {})


def _entry_from_pem(pem_entry: dict) -> dict:
    """Convert per_experiment_metrics[batch] → {overall, spike, flat}."""
    return {
        "overall": _first_target(pem_entry.get("overall", {})),
        "spike":   _first_target(pem_entry.get("spike", {})),
        "flat":    _first_target(pem_entry.get("non_spike", {})),
    }


def _entry_from_toplevel(d: dict) -> dict:
    """Convert top-level metrics/metrics_spike/metrics_flat → {overall, spike, flat}."""
    return {
        "overall": _first_target(d.get("metrics", {})),
        "spike":   _first_target(d.get("metrics_spike", {})),
        "flat":    _first_target(d.get("metrics_flat", {})),
    }


def _load_json(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------

def collect_all() -> dict:
    """Return data[config][batch] = {overall, spike, flat} dicts."""
    data: dict = {}

    # ── per-batch folders: absolute, anchor, one_by_one ────────────────────
    for config in ("absolute", "anchor", "one_by_one"):
        folder = BASE / config
        if not folder.exists():
            print(f"  WARNING: {folder} not found — skipping")
            continue
        for batch_dir in sorted(folder.iterdir()):
            if not batch_dir.is_dir():
                continue
            batch = batch_dir.name
            rj = batch_dir / "results.json"
            if not rj.exists():
                continue
            d = _load_json(rj)
            pem = d.get("per_experiment_metrics", {})
            if batch in pem:
                entry = _entry_from_pem(pem[batch])
            else:
                entry = _entry_from_toplevel(d)
            data.setdefault(config, {})[batch] = entry
        found = sorted(data.get(config, {}).keys())
        print(f"  {config:<15}: {found}")

    # ── LOO ─────────────────────────────────────────────────────────────────
    loo_folder = BASE / "LOO"
    if loo_folder.exists():
        for loo_dir in sorted(loo_folder.iterdir()):
            if not loo_dir.is_dir():
                continue
            rj = loo_dir / "results.json"
            if not rj.exists():
                continue
            d = _load_json(rj)
            pem_keys = list(d.get("per_experiment_metrics", {}).keys())
            if not pem_keys:
                continue
            batch = pem_keys[0]
            entry = _entry_from_toplevel(d)
            data.setdefault("LOO", {})[batch] = entry
        found = sorted(data.get("LOO", {}).keys())
        print(f"  {'LOO':<15}: {found}")
    else:
        print(f"  WARNING: {loo_folder} not found — skipping")

    # ── finetune ─────────────────────────────────────────────────────────────
    ft_folder = BASE / "finetune"
    if ft_folder.exists():
        for test_dir in sorted(ft_folder.iterdir()):
            if not test_dir.is_dir():
                continue
            for sub_dir in sorted(test_dir.iterdir()):
                if not sub_dir.is_dir():
                    continue
                if not sub_dir.name.startswith("finetune_"):
                    continue
                rj = sub_dir / "results.json"
                if not rj.exists():
                    continue
                d = _load_json(rj)
                # batch name: finetune_KAU079 → KAU079
                batch = sub_dir.name[len("finetune_"):]
                pem = d.get("per_experiment_metrics", {})
                if batch in pem:
                    entry = _entry_from_pem(pem[batch])
                else:
                    entry = _entry_from_toplevel(d)
                data.setdefault("finetune", {})[batch] = entry
        found = sorted(data.get("finetune", {}).keys())
        print(f"  {'finetune':<15}: {found}")
    else:
        print(f"  WARNING: {ft_folder} not found — skipping")

    # ── multi-batch concat files ─────────────────────────────────────────────
    concat_specs = {
        "concat_4batch":    "KAU084_KAU081_KAU071_KAU079",
        "no_token_7batch":  "no_tokenKAU084_KAU081_KAU071_KAU074_KAU075_KAU076_KAU079",
        "token_7batch":     "token_KAU084_KAU081_KAU071_KAU074_KAU075_KAU076_KAU079",
    }
    for config, folder_name in concat_specs.items():
        rj = BASE / folder_name / "results.json"
        if not rj.exists():
            print(f"  WARNING: {rj} not found — skipping {config}")
            continue
        d = _load_json(rj)
        pem = d.get("per_experiment_metrics", {})
        for batch, pem_entry in pem.items():
            data.setdefault(config, {})[batch] = _entry_from_pem(pem_entry)
        found = sorted(data.get(config, {}).keys())
        print(f"  {config:<15}: {found}")

    return data


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def _style_cell(cell, fill=None, font=None, bold=False, align="center"):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    elif bold:
        cell.font = FONT_BOLD
    cell.alignment = Alignment(horizontal=align, vertical="center")


def write_excel(data: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n_configs = len(CONFIG_ORDER)
    total_cols = 1 + n_configs  # batch col + one per config

    sections = [
        ("overall", "Overall"),
        ("spike",   "Spike"),
        ("flat",    "Non-Spike"),
    ]

    for metric in METRICS:
        sheet_title = METRIC_LABELS[metric]
        ws = wb.create_sheet(title=sheet_title)
        row = 1

        for section_key, section_label in sections:
            # ── Section title ───────────────────────────────────────────────
            cell = ws.cell(row, 1, section_label)
            _style_cell(cell, fill=FILL_SECTION, font=FONT_WHITE_BOLD, align="left")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=total_cols)
            row += 1

            # ── Column headers ──────────────────────────────────────────────
            cell = ws.cell(row, 1, "Batch")
            _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            for ci, config in enumerate(CONFIG_ORDER, start=2):
                cell = ws.cell(row, ci, config)
                _style_cell(cell, fill=FILL_HEADER, font=FONT_WHITE_BOLD)
            row += 1

            # ── Data rows ───────────────────────────────────────────────────
            active_batches = [
                b for b in BATCH_ORDER
                if any(
                    data.get(c, {}).get(b, {}).get(section_key, {}).get(metric) is not None
                    for c in CONFIG_ORDER
                )
            ]

            for ri, batch in enumerate(active_batches):
                fill = FILL_ROW_A if ri % 2 == 0 else FILL_ROW_B
                cell = ws.cell(row, 1, batch)
                _style_cell(cell, fill=fill, bold=True, align="left")
                for ci, config in enumerate(CONFIG_ORDER, start=2):
                    val = (data.get(config, {})
                               .get(batch, {})
                               .get(section_key, {})
                               .get(metric))
                    cell = ws.cell(row, ci)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val is not None:
                        cell.value = round(val, 6)
                        cell.number_format = "0.000000"
                row += 1

            row += 1  # blank row between sections

        # ── Column widths ───────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 12
        for ci, config in enumerate(CONFIG_ORDER, start=2):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(config) + 2, 14)

    wb.save(OUT)
    print(f"\nWritten: {OUT}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"\nCollecting from: {BASE}")
    print("-" * 60)
    data = collect_all()

    configs_found = [c for c in CONFIG_ORDER if c in data]
    print(f"\nConfigs collected: {configs_found}")
    print(f"Writing {len(METRICS)} sheets × {len(sections_summary(data))} active batches ...")
    write_excel(data)


def sections_summary(data):
    batches = set()
    for config_data in data.values():
        batches.update(config_data.keys())
    return batches


if __name__ == "__main__":
    main()
